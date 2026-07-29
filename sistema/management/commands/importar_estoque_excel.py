from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from sistema.models import ItemEstoque, MovimentacaoEstoque


ARQUIVOS_PADRAO = [
    "COPA ABRIL 2026 - Copia 4.xlsx",
    "MATERIAL DE EXPEDIENTE MAIO.xlsx",
]


def texto_limpo(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def texto_chave(valor):
    texto = texto_limpo(valor).upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.replace(" ", "").replace("_", "")
    return texto


def decimal_positivo(valor):
    if valor in (None, ""):
        return None
    try:
        numero = Decimal(str(valor).replace(",", "."))
    except (InvalidOperation, ValueError):
        return None
    if numero < 0:
        return None
    return numero


def decimal_moeda(valor):
    numero = decimal_positivo(valor)
    if numero is None:
        return None
    return numero.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def inteiro_positivo(valor):
    if valor in (None, ""):
        return None
    try:
        numero = int(float(valor))
    except (TypeError, ValueError):
        return None
    if numero < 0:
        return None
    return numero


def data_para_date(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return None


class Command(BaseCommand):
    help = "Importa itens e movimentacoes de estoque a partir de planilhas Excel."

    def add_arguments(self, parser):
        parser.add_argument(
            "--arquivo",
            action="append",
            dest="arquivos",
            help="Caminho completo de arquivo .xlsx. Pode ser informado varias vezes.",
        )
        parser.add_argument(
            "--downloads-dir",
            default=str(Path.home() / "Downloads"),
            help="Diretorio usado para localizar os arquivos padrao.",
        )
        parser.add_argument(
            "--somente-cadastro",
            action="store_true",
            help="Importa somente a aba PROD (sem ENTRADAS e SAIDAS).",
        )

    def handle(self, *args, **options):
        arquivos = options.get("arquivos") or []
        downloads_dir = Path(options["downloads_dir"])

        if not arquivos:
            for nome in ARQUIVOS_PADRAO:
                caminho = downloads_dir / nome
                if caminho.exists():
                    arquivos.append(str(caminho))

        if not arquivos:
            raise CommandError("Nenhum arquivo informado ou encontrado para importar.")

        resumo_total = {
            "itens_criados": 0,
            "itens_atualizados": 0,
            "entradas_criadas": 0,
            "saidas_criadas": 0,
            "movimentos_ignorados": 0,
        }

        for arquivo in arquivos:
            caminho = Path(arquivo)
            if not caminho.exists():
                raise CommandError(f"Arquivo nao encontrado: {caminho}")

            categoria = (
                ItemEstoque.Categoria.COPA
                if "copa" in caminho.name.lower()
                else ItemEstoque.Categoria.EXPEDIENTE
            )
            resumo = self.importar_arquivo(
                caminho=caminho,
                categoria=categoria,
                somente_cadastro=options["somente_cadastro"],
            )
            for chave in resumo_total:
                resumo_total[chave] += resumo[chave]

            self.stdout.write(
                self.style.SUCCESS(
                    f"[{caminho.name}] itens+={resumo['itens_criados']}/{resumo['itens_atualizados']} "
                    f"entradas+={resumo['entradas_criadas']} saidas+={resumo['saidas_criadas']} "
                    f"ignorados={resumo['movimentos_ignorados']}"
                )
            )

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("Resumo final da importacao:"))
        for chave, valor in resumo_total.items():
            self.stdout.write(f" - {chave}: {valor}")

    @transaction.atomic
    def importar_arquivo(self, caminho, categoria, somente_cadastro=False):
        wb = load_workbook(filename=caminho, data_only=True)
        sheet_prod = wb["PROD"] if "PROD" in wb.sheetnames else None
        sheet_entradas = wb["ENTRADAS"] if "ENTRADAS" in wb.sheetnames else None
        sheet_saidas = wb["SAIDAS"] if "SAIDAS" in wb.sheetnames else None

        if not sheet_prod:
            raise CommandError(f"Aba PROD nao encontrada em: {caminho.name}")

        resumo = {
            "itens_criados": 0,
            "itens_atualizados": 0,
            "entradas_criadas": 0,
            "saidas_criadas": 0,
            "movimentos_ignorados": 0,
        }

        itens_por_nome = {}
        colunas_prod = self.mapear_colunas(sheet_prod, 6)
        col_nome = colunas_prod.get("CADASTRODEPRODUTOS") or colunas_prod.get("ITEM")
        col_codigo = colunas_prod.get("CODIGOPROPRIO")
        col_unidade = colunas_prod.get("UNIDADEDEMEDIDA")
        col_minimo = colunas_prod.get("ESTOQUEMINIMO")
        col_custo = colunas_prod.get("CUSTOUNITARIO")

        if not col_nome:
            raise CommandError(f"Nao foi possivel mapear colunas da aba PROD em {caminho.name}.")

        for row_idx in range(7, sheet_prod.max_row + 1):
            nome = texto_limpo(sheet_prod.cell(row=row_idx, column=col_nome).value)
            if not nome:
                continue

            codigo = (
                texto_limpo(sheet_prod.cell(row=row_idx, column=col_codigo).value)
                if col_codigo
                else ""
            )
            unidade = (
                texto_limpo(sheet_prod.cell(row=row_idx, column=col_unidade).value)
                if col_unidade
                else "un"
            ) or "un"
            estoque_minimo = (
                inteiro_positivo(sheet_prod.cell(row=row_idx, column=col_minimo).value) if col_minimo else 0
            ) or 0
            custo_unitario = (
                decimal_moeda(sheet_prod.cell(row=row_idx, column=col_custo).value) if col_custo else Decimal("0")
            ) or Decimal("0")

            item, criado = ItemEstoque.objects.get_or_create(
                area=ItemEstoque.Area.ADMINISTRATIVO,
                nome=nome,
                defaults={
                    "categoria": categoria,
                    "codigo_proprio": codigo,
                    "unidade_medida": unidade,
                    "estoque_minimo": estoque_minimo,
                    "custo_unitario": custo_unitario,
                    "ativo": True,
                },
            )
            if criado:
                resumo["itens_criados"] += 1
            else:
                atualizado = False
                if (
                    item.categoria != categoria
                    and item.categoria != ItemEstoque.Categoria.GERAL
                    and categoria != ItemEstoque.Categoria.GERAL
                ):
                    item.categoria = ItemEstoque.Categoria.GERAL
                    atualizado = True
                if codigo and not item.codigo_proprio:
                    item.codigo_proprio = codigo
                    atualizado = True
                if (
                    unidade
                    and item.unidade_medida in {"", "un", "UN", "UNIDADE", "UNIDADES"}
                    and item.unidade_medida != unidade
                ):
                    item.unidade_medida = unidade
                    atualizado = True
                if estoque_minimo > 0 and item.estoque_minimo == 0:
                    item.estoque_minimo = estoque_minimo
                    atualizado = True
                if custo_unitario and item.custo_unitario == 0:
                    item.custo_unitario = custo_unitario
                    atualizado = True
                if atualizado:
                    item.save()
                    resumo["itens_atualizados"] += 1

            itens_por_nome[nome.upper()] = item

        if somente_cadastro:
            return resumo

        if sheet_entradas:
            colunas_entradas = self.mapear_colunas(sheet_entradas, 6)
            col_data = colunas_entradas.get("DATADACOMPRA")
            col_nome = colunas_entradas.get("PRODUTO")
            col_qtd = colunas_entradas.get("QUANTIDADE")
            col_custo = colunas_entradas.get("CUSTOUNITARIO")
            if not col_nome or not col_qtd:
                raise CommandError(
                    f"Nao foi possivel mapear colunas da aba ENTRADAS em {caminho.name}."
                )
            for row_idx in range(7, sheet_entradas.max_row + 1):
                origem_externa = f"{caminho.name}|ENTRADAS|{row_idx}"
                if MovimentacaoEstoque.objects.filter(origem_externa=origem_externa).exists():
                    resumo["movimentos_ignorados"] += 1
                    continue

                data_mov = (
                    data_para_date(sheet_entradas.cell(row=row_idx, column=col_data).value)
                    if col_data
                    else None
                ) or date.today()
                nome = texto_limpo(sheet_entradas.cell(row=row_idx, column=col_nome).value)
                quantidade = inteiro_positivo(sheet_entradas.cell(row=row_idx, column=col_qtd).value)
                custo = (
                    decimal_moeda(sheet_entradas.cell(row=row_idx, column=col_custo).value)
                    if col_custo
                    else None
                )

                if not nome or not quantidade or quantidade <= 0:
                    continue

                item = itens_por_nome.get(nome.upper())
                if not item:
                    item = ItemEstoque.objects.filter(
                        area=ItemEstoque.Area.ADMINISTRATIVO, nome__iexact=nome
                    ).first()
                if not item:
                    item = ItemEstoque.objects.create(
                        area=ItemEstoque.Area.ADMINISTRATIVO,
                        categoria=categoria,
                        nome=nome,
                        unidade_medida="un",
                        estoque_minimo=0,
                        custo_unitario=custo or Decimal("0"),
                    )
                    itens_por_nome[nome.upper()] = item
                    resumo["itens_criados"] += 1

                MovimentacaoEstoque.objects.create(
                    item=item,
                    data_movimentacao=data_mov,
                    tipo=MovimentacaoEstoque.TipoMovimento.ENTRADA,
                    quantidade=quantidade,
                    quantidade_devolvida=0,
                    custo_unitario=custo or item.custo_unitario or Decimal("0"),
                    observacao=f"Importado de {caminho.name} - ENTRADAS",
                    origem_externa=origem_externa,
                )
                resumo["entradas_criadas"] += 1

        if sheet_saidas:
            colunas_saidas = self.mapear_colunas(sheet_saidas, 6)
            col_data = colunas_saidas.get("DATADASAIDA")
            col_nome = colunas_saidas.get("PRODUTO")
            col_qtd = colunas_saidas.get("QUANTIDADE")
            col_devolvida = colunas_saidas.get("QUANTDEVOLVIDA")
            col_custo = colunas_saidas.get("PRECOUNITARIO")
            col_setor = colunas_saidas.get("SETOR")
            if not col_nome or not col_qtd:
                raise CommandError(f"Nao foi possivel mapear colunas da aba SAIDAS em {caminho.name}.")
            for row_idx in range(7, sheet_saidas.max_row + 1):
                origem_externa = f"{caminho.name}|SAIDAS|{row_idx}"
                if MovimentacaoEstoque.objects.filter(origem_externa=origem_externa).exists():
                    resumo["movimentos_ignorados"] += 1
                    continue

                nome = texto_limpo(sheet_saidas.cell(row=row_idx, column=col_nome).value)
                if not nome:
                    continue

                quantidade = inteiro_positivo(sheet_saidas.cell(row=row_idx, column=col_qtd).value)
                if not quantidade or quantidade <= 0:
                    continue

                col_d = sheet_saidas.cell(row=row_idx, column=col_devolvida).value if col_devolvida else None
                col_e = sheet_saidas.cell(row=row_idx, column=col_custo).value if col_custo else None
                devolvida = inteiro_positivo(col_d) or 0
                custo = decimal_moeda(col_e)

                d_num = decimal_moeda(col_d)
                e_num = decimal_moeda(col_e)
                if d_num and (not e_num or e_num <= quantidade) and d_num > quantidade * 2:
                    custo = d_num
                    devolvida = 0

                if devolvida > quantidade:
                    devolvida = quantidade

                setor = (
                    texto_limpo(sheet_saidas.cell(row=row_idx, column=col_setor).value)
                    if col_setor
                    else ""
                )
                data_mov = (
                    data_para_date(sheet_saidas.cell(row=row_idx, column=col_data).value)
                    if col_data
                    else None
                ) or date.today()

                item = itens_por_nome.get(nome.upper())
                if not item:
                    item = ItemEstoque.objects.filter(
                        area=ItemEstoque.Area.ADMINISTRATIVO, nome__iexact=nome
                    ).first()
                if not item:
                    item = ItemEstoque.objects.create(
                        area=ItemEstoque.Area.ADMINISTRATIVO,
                        categoria=categoria,
                        nome=nome,
                        unidade_medida="un",
                        estoque_minimo=0,
                        custo_unitario=custo or Decimal("0"),
                    )
                    itens_por_nome[nome.upper()] = item
                    resumo["itens_criados"] += 1

                saldo_necessario = quantidade - devolvida
                if saldo_necessario > item.quantidade_atual:
                    resumo["movimentos_ignorados"] += 1
                    continue

                MovimentacaoEstoque.objects.create(
                    item=item,
                    data_movimentacao=data_mov,
                    tipo=MovimentacaoEstoque.TipoMovimento.SAIDA,
                    quantidade=quantidade,
                    quantidade_devolvida=devolvida,
                    custo_unitario=custo or item.custo_unitario or Decimal("0"),
                    setor=setor[:80],
                    observacao=f"Importado de {caminho.name} - SAIDAS",
                    origem_externa=origem_externa,
                )
                resumo["saidas_criadas"] += 1

        return resumo

    def mapear_colunas(self, sheet, linha_cabecalho):
        colunas = {}
        for col_idx in range(1, sheet.max_column + 1):
            chave = texto_chave(sheet.cell(row=linha_cabecalho, column=col_idx).value)
            if chave:
                colunas[chave] = col_idx
        return colunas
